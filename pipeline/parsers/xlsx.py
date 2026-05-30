"""
XLSX parser using openpyxl.
Each sheet becomes a separate page entry.
Rows are joined as tab-separated text.
"""

from pathlib import Path


def extract(filepath: str) -> list[dict]:
    """
    Returns a list of sheet dicts: {"page": sheet_index, "text": str, "sheet": sheet_name}
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(filepath)

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheets = []

    for i, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                rows.append(row_text)
        if rows:
            sheets.append({"page": i, "sheet": sheet_name,
                          "text": "\n".join(rows)})

    wb.close()
    return sheets
